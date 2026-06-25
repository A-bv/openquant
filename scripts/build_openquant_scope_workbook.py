from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUT = Path("docs/openquant_scope_table.xlsx")


COURSE_OBJECTIVE = (
    "This course introduces students to financial valuation and value enhancement "
    "through financial decision making."
)


SOURCE_FILES = [
    ("Course notes", "EPFL_Finance_Complete_Notes.docx", "Course structure, concepts, worked solution summaries"),
    ("Formula sheet", "z Formula Sheet Final 2011 - copie.pdf", "Official formula reference"),
    ("Workbook", "EPFL_Finance_Master_Workbook.xlsx", "Interactive course formula sheets and examples"),
    ("Sample exam", "SampleExam1.pdf", "Final exam problems and solutions"),
    ("Sample exam", "SampleExam2.pdf", "Midterm B problems and solutions"),
    ("Sample exam", "SampleExam3.pdf", "Midterm problems and solutions"),
    ("OpenQuant repo", "core/, api/, frontend/, backtest/, tests/", "Implemented product, formulas, tests, and backtests"),
]


WORKBOOK_SHEETS = [
    ("Cover", 0, "Course metadata and workbook navigation"),
    ("TVM & Bonds", 59, "PV/FV, perpetuity, growing perpetuity, annuity, bond pricing, YTM, stock valuation, NWC mini-block"),
    ("Rate & Spot Rates", 25, "Rate conversion, zero-coupon spot rates, forward rates, coupon bond from spots"),
    ("Duration & PI", 45, "Macaulay duration, modified duration, profitability index, NPV vs IRR"),
    ("Stats Basics", 27, "Expected return, variance, standard deviation, covariance, correlation"),
    ("Portfolio Theory", 20, "Portfolio weights, expected return, covariance matrix, portfolio variance and volatility"),
    ("Adv Portfolio", 17, "Minimum variance weight, idiosyncratic variance, capital gain rate"),
    ("CAPM", 11, "Beta, CAPM required return, security market line, alpha"),
    ("WACC", 7, "Cost of equity, cost of debt, WACC, Modigliani-Miller intuition"),
    ("APV & Distress", 32, "APV, PV tax shield, bankruptcy costs, trade-off theory"),
    ("DCF Valuation", 50, "FCF projections, terminal value, enterprise value, equity bridge, scenario valuation"),
    ("Disney DCF", 32, "Damodaran-style 10-year DCF case study"),
    ("WB Intrinsic Value", 18, "Buffett/BVPS intrinsic value method"),
    ("Derivatives Calc", 97, "Forward/futures pricing, option payoffs, put-call parity, binomial, Black-Scholes"),
    ("TOTAL", 440, "Total formula cells across the EPFL master workbook"),
]


CONCEPT_ROWS = [
    ("Course identity", "Principles of Finance, likely based on Berk & DeMarzo Corporate Finance, 2nd ed.", "Course anchor", "Notes / workbook", "No", "No", "Reference", "No", "No", "Reference row only"),
    ("Official objective", "Financial valuation and value enhancement through financial decision making", "Final course objective", "Official course slide/objective", "Yes", "Yes", "Partial", "Partial", "Yes", "OpenQuant covers the listed-company valuation subset today"),
    ("Product objective", "Apply the course objective to real listed-company data", "OpenQuant goal", "README / strategy", "No", "No", "Yes", "Yes", "Yes", "This is the current product spine"),
    ("Authenticity layer", "Backtest past OpenQuant conclusions against later historical data", "Product validation", "backtest/ + tests", "No", "No", "Yes", "Yes", "Yes", "Backtests validate informativeness, not formula correctness"),
    ("Time value", "Present value / future value / discounting", "Foundation for valuation", "Formula sheet, workbook, exams", "Yes", "Yes", "Yes", "Yes", "No", "Used behind DCF; no standalone TVM lab"),
    ("Time value", "Perpetuity and growing perpetuity", "Terminal value and stock valuation", "Formula sheet, workbook, exams", "Yes", "Yes", "Yes", "Yes", "No", "Used directly in terminal value and DDM tests"),
    ("Time value", "Annuity and growing annuity", "Multi-period cash-flow valuation", "Formula sheet, workbook, exams", "Yes", "Yes", "Partial", "No", "Later", "Growing annuity is implemented/tested; plain annuity is not a reusable function"),
    ("Rates", "Rate conversion and compounding frequency", "Equivalent rates and discounting", "Formula sheet, workbook, exams", "Yes", "Yes", "No", "No", "Later", "Useful for Bonds/Rates Lab, not current stock product"),
    ("Rates", "Spot rates and forward rates", "No-arbitrage rates", "Formula sheet, workbook, exams", "Yes", "Yes", "No", "No", "Later", "Bonds/Rates Lab"),
    ("Bonds", "Bond pricing and yield to maturity", "Debt valuation", "Formula sheet, workbook, exams", "Yes", "Yes", "No", "No", "Later", "Could appear as demonstration, not core stock product"),
    ("Bonds", "Duration and price sensitivity", "Interest-rate risk", "Formula sheet, workbook", "No", "Yes", "No", "No", "Later", "Bonds/Rates Lab"),
    ("Statistics", "Expected return, variance, standard deviation", "Risk measurement", "Formula sheet, workbook, exams", "Yes", "Yes", "Partial", "No", "Later", "Volatility helpers exist; expected return/variance are test calculations, not reusable product functions"),
    ("Statistics", "Covariance and correlation", "Co-movement and diversification", "Formula sheet, workbook, exams", "Yes", "Yes", "Partial", "No", "Later", "Covariance is used for beta; correlation is only an input helper/test item"),
    ("Portfolio", "Portfolio expected return and variance", "Diversification and portfolio risk", "Formula sheet, workbook, exams", "Yes", "Yes", "Partial", "No", "Later", "Some formulas are test calculations; no portfolio product module exists"),
    ("Portfolio", "Minimum variance portfolio", "Optimization under risk", "Workbook, exams", "Yes", "Yes", "Yes", "No", "Later", "Implemented/tested; no UI"),
    ("Portfolio", "Efficient frontier / CML / Sharpe", "Risk-adjusted allocation", "Notes, formula sheet, workbook, exams", "Yes", "Yes", "Partial", "No", "Later", "Sharpe implemented; CML/frontier are not implemented/productized"),
    ("CAPM", "Beta and systematic risk", "Required return under market risk", "Formula sheet, workbook, exams", "Yes", "Yes", "Yes", "Yes", "Yes", "Used in WACC and current stock valuation"),
    ("CAPM", "CAPM required return", "Cost of equity", "Formula sheet, workbook, exams", "Yes", "Yes", "Yes", "Yes", "Yes", "Core current product input"),
    ("CAPM", "Alpha and SML interpretation", "Performance vs required return", "Notes, workbook", "Maybe", "Yes", "No", "No", "Later", "Useful for Portfolio Lab, not current flow"),
    ("Capital structure", "WACC", "Discount rate for enterprise cash flows", "Formula sheet, workbook, exams", "Yes", "Yes", "Yes", "Yes", "Yes", "Core current product input"),
    ("Capital structure", "Unlever/relever beta", "Adjust risk for leverage", "Formula sheet, workbook, exams", "Yes", "Yes", "Partial", "No", "Later", "Unlevering is implemented/tested; relevering is not separately implemented or visible"),
    ("Capital structure", "Modigliani-Miller propositions", "Capital-structure policy reasoning", "Formula sheet, workbook, exams", "Yes", "Yes", "No", "No", "Later", "Theory in course/exams, but not an OpenQuant product flow"),
    ("Capital structure", "APV and PV tax shield", "Separate operating value from financing effects", "Formula sheet, workbook, exams", "Yes", "Yes", "Partial", "No", "Later", "PVTS helper exists; APV not productized"),
    ("Capital structure", "Bankruptcy / distress trade-off", "Limits of debt financing", "Formula sheet, workbook, exams", "Yes", "Yes", "No", "No", "Later", "Needed for full capital-structure lab"),
    ("Cash flows", "Free cash flow from operating components", "Value creation measurement", "Workbook, exams", "Yes", "Yes", "Yes", "Yes", "Yes", "Core current product"),
    ("Cash flows", "Net working capital and change in NWC", "Cash-flow timing and investment needs", "Workbook, exams", "Yes", "Yes", "Partial", "No", "Later", "Used in tests/project reasoning; not exposed in product"),
    ("Valuation", "DCF enterprise value and equity value bridge", "Company valuation", "Formula sheet, workbook, exams", "Yes", "Yes", "Yes", "Yes", "Yes", "Core current product"),
    ("Valuation", "Terminal value", "Long-run value assumption", "Formula sheet, workbook, exams", "Yes", "Yes", "Yes", "Yes", "Yes", "Core current product and key risk"),
    ("Valuation", "Reverse DCF / market-implied growth", "Translate price into assumptions", "OpenQuant product extension", "No", "No", "Yes", "Yes", "Yes", "Central OpenQuant product contribution"),
    ("Valuation", "Sensitivity analysis", "How conclusions change with assumptions", "Workbook / product", "No", "Yes", "Yes", "Yes", "Yes", "Productized as growth x WACC heatmap"),
    ("Valuation", "Multiples sanity check", "Cross-check DCF against market measures", "Product", "No", "No", "Yes", "Yes", "Yes", "Productized, not a core course formula"),
    ("Investment decisions", "NPV vs IRR and scale problem", "Choose value-creating projects", "Formula sheet, workbook, exams", "Yes", "Yes", "Yes", "No", "Later", "Implemented/tested; Project Decision Lab not productized"),
    ("Stocks", "Dividend discount model", "Equity valuation from dividends", "Formula sheet, workbook, exams", "Yes", "Yes", "Yes", "No", "Later", "Implemented/tested but current product uses FCF DCF"),
    ("Stocks", "Buffett/BVPS intrinsic value", "Alternative valuation demonstration", "Workbook", "No", "Yes", "No", "No", "Later", "Can be demo only; not course core objective"),
    ("Derivatives", "Forwards, futures, swaps", "Derivative securities and corporate finance use", "Formula sheet, workbook, exams", "Yes", "Yes", "No", "No", "Only if scope expands", "Confirmed out of current OpenQuant scope"),
    ("Derivatives", "Options, put-call parity, binomial, Black-Scholes", "Derivative pricing and payoff logic", "Formula sheet, workbook, exams", "Yes", "Yes", "No", "No", "Only if scope expands", "Confirmed out of current OpenQuant scope"),
]


FORMULA_ROWS = [
    ("Time value", "Present value of cash flows", "PV = sum(C_t / (1+r)^t)", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "No", "DCF discounting"),
    ("Time value", "Discount factor", "DF_t = 1 / (1+r)^t", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "No", "Used internally"),
    ("Time value", "Future value factor", "FV = PV x (1+r)^t", "Yes", "Yes", "Yes", "No", "No", "No", "No", "No current product need"),
    ("Time value", "Perpetuity", "PV = C / r", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "No", "Special case of growing perpetuity"),
    ("Time value", "Growing perpetuity", "PV = C_1 / (r-g)", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "No", "Terminal value and DDM"),
    ("Time value", "Annuity factor", "PV = C/r x [1 - 1/(1+r)^N]", "Yes", "Yes", "Yes", "No", "No", "No", "Later", "Present in course/workbook; no reusable OpenQuant annuity function"),
    ("Time value", "Growing annuity", "PV = C_1/(r-g) x [1 - ((1+g)/(1+r))^N]", "Yes", "Yes", "Yes", "Yes", "No", "No", "Later", "Sample exam tested"),
    ("Rates", "Compounding conversion", "(1+r_A)^(p_A) = (1+r_B)^(p_B)", "Yes", "Yes", "Yes", "No", "No", "No", "Later", "Bonds/Rates Lab"),
    ("Rates", "Continuous compounding", "FV = PV x exp(r x t)", "Yes", "Yes", "Yes", "No", "No", "No", "Later", "Mainly rates/derivatives"),
    ("Rates", "Spot rate from zero-coupon bond", "r_0,t = (FV/P)^(1/t) - 1", "Yes", "Yes", "Yes", "No", "No", "No", "Later", "Bonds/Rates Lab"),
    ("Rates", "Forward rate from spot rates", "f_m,n = [((1+r_n)^n)/((1+r_m)^m)]^(1/(n-m)) - 1", "Yes", "Yes", "Yes", "No", "No", "No", "Later", "Bonds/Rates Lab"),
    ("Bonds", "Coupon bond price", "P = sum(CPN_t / (1+r_t)^t) + Face/(1+r_T)^T", "Yes", "Yes", "Yes", "No", "No", "No", "Later", "Bonds/Rates Lab"),
    ("Bonds", "Yield to maturity", "Price = sum(CF_t / (1+YTM)^t)", "Yes", "Yes", "Yes", "No", "No", "No", "Later", "Requires solver if productized"),
    ("Bonds", "Macaulay duration", "D = sum(t x PV(CF_t)) / Price", "Yes", "Yes", "Yes", "No", "No", "No", "Later", "Bonds/Rates Lab"),
    ("Bonds", "Modified duration", "D_mod = D_mac / (1+y/k)", "Yes", "Yes", "Yes", "No", "No", "No", "Later", "Bonds/Rates Lab"),
    ("Statistics", "Expected return", "E[R] = sum(p_i x R_i)", "Yes", "Yes", "Yes", "No", "No", "No", "Later", "Verified in tests with pandas/numpy, but no reusable OpenQuant function"),
    ("Statistics", "Variance", "Var(R) = sum(p_i x (R_i - E[R])^2)", "Yes", "Yes", "Yes", "No", "No", "No", "Later", "Verified in tests with pandas/numpy, but no reusable OpenQuant function"),
    ("Statistics", "Standard deviation", "SD(R) = sqrt(Var(R))", "Yes", "Yes", "Yes", "Yes", "No", "No", "Later", "Volatility annualisation exists in core.common; no standalone probability-law module"),
    ("Statistics", "Covariance", "Cov(X,Y) = E[(X-E[X])(Y-E[Y])]", "Yes", "Yes", "Yes", "Yes", "No", "Yes", "Later", "Used in beta estimation via Cov/Var, not a standalone product feature"),
    ("Statistics", "Correlation", "rho_XY = Cov(X,Y)/(sigma_X sigma_Y)", "Yes", "Yes", "Yes", "No", "No", "No", "Later", "Used as an input in beta_from_correlation; no correlation-computation function"),
    ("Risk", "Beta from covariance", "beta_i = Cov(R_i,R_M) / Var(R_M)", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Current product estimates beta from market price data"),
    ("Risk", "Beta from correlation", "beta_i = rho_iM x sigma_i / sigma_M", "Yes", "No", "Yes", "Yes", "No", "No", "Later", "Exam 2 test; not current product input method"),
    ("Portfolio", "Portfolio return", "E[R_p] = sum(w_i x E[R_i])", "Yes", "Yes", "Yes", "No", "No", "No", "Later", "Computed directly in tests; no reusable product function"),
    ("Portfolio", "Two-asset variance", "Var_p = w_A^2 sigma_A^2 + w_B^2 sigma_B^2 + 2w_Aw_Brho_ABsigma_Asigma_B", "Yes", "Yes", "Yes", "No", "No", "No", "Later", "Computed directly in tests; min-variance helper exists separately"),
    ("Portfolio", "N-asset portfolio beta", "beta_p = sum(w_i x beta_i)", "Yes", "No", "No", "No", "No", "No", "Later", "Formula sheet line; add in Portfolio Lab"),
    ("Portfolio", "Minimum variance weight", "w_A* = (sigma_B^2 - rho sigma_A sigma_B)/(sigma_A^2 + sigma_B^2 - 2rho sigma_A sigma_B)", "No", "Yes", "Yes", "Yes", "No", "No", "Later", "Implemented/tested; no product UI"),
    ("Portfolio", "Sharpe ratio", "Sharpe = (E[R] - R_f) / sigma", "No", "Yes", "Yes", "Yes", "No", "No", "Later", "Implemented/tested; Portfolio Lab"),
    ("CAPM", "CAPM required return", "r_i = r_f + beta_i x (E[R_M] - r_f)", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Current product WACC cost of equity"),
    ("CAPM", "Alpha", "alpha_i = R_i - [r_f + beta_i x (E[R_M]-r_f)]", "No", "Yes", "Yes", "No", "No", "No", "Later", "No standalone alpha implementation or product flow"),
    ("WACC", "WACC", "WACC = E/(D+E) x r_E + D/(D+E) x r_D x (1-T_c)", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Core current product"),
    ("Capital structure", "Hamada unlever beta", "beta_U = beta_E / [1 + (1-T_c)D/E]", "Yes", "Yes", "Yes", "Yes", "No", "Yes", "Later", "Implemented/tested, but not visible or used in current WACC output"),
    ("Capital structure", "Relever beta", "beta_E = beta_U x [1 + (1-T_c)D/E]", "Yes", "Yes", "Yes", "No", "No", "Yes", "Later", "Unlevering exists; full explicit relever-beta formula is not a separate implementation"),
    ("Capital structure", "MM I no taxes", "V_L = V_U", "Yes", "Yes", "Yes", "No", "No", "No", "Later", "Theory in course/exams, not an OpenQuant formula implementation"),
    ("Capital structure", "MM II no taxes", "r_E = r_A + (D/E)(r_A-r_D)", "Yes", "Yes", "Yes", "No", "No", "No", "Later", "Capital Structure Lab"),
    ("Capital structure", "APV", "APV = base-case NPV + NPV(financing side effects)", "Yes", "Yes", "Yes", "No", "No", "No", "Later", "PVTS exists, but full APV formula/workflow is not implemented"),
    ("Capital structure", "PV tax shield", "PVTS = sum(T_c x r_D x D_t / (1+r_D)^t)", "Yes", "Yes", "Yes", "Yes", "No", "No", "Later", "Exam tested; not current UI"),
    ("Capital structure", "Trade-off value", "V_L = V_U + PVTS - PV distress costs", "Yes", "Yes", "Yes", "No", "No", "No", "Later", "Needed for full financing policy"),
    ("Cash flows", "Free cash flow", "FCF = EBIT(1-T_c) + Depreciation - Capex - Delta NWC", "No", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Core current product"),
    ("Cash flows", "Net working capital", "NWC = AR + Inventory - AP", "No", "Yes", "Yes", "No", "No", "Yes", "Later", "Change in working capital can be used as input; explicit NWC component formula is not implemented"),
    ("Valuation", "Terminal value", "TV_n = FCF_n x (1+g) / (WACC-g)", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Core current product"),
    ("Valuation", "Enterprise value DCF", "EV = sum(FCF_t/(1+WACC)^t) + TV/(1+WACC)^n", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Core current product"),
    ("Valuation", "Equity value bridge", "Equity value = EV - Net debt", "No", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Core current product"),
    ("Valuation", "Intrinsic value per share", "IV/share = Equity value / diluted shares", "No", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Core current product"),
    ("Valuation", "Reverse DCF", "Solve for g such that IV(g) = current market price", "No", "No", "No", "Yes", "Yes", "Yes", "Yes", "OpenQuant-specific practical layer"),
    ("Investment decisions", "NPV", "NPV = sum(C_t/(1+r)^t)", "Yes", "Yes", "Yes", "Yes", "No", "Yes", "Later", "Implemented/tested; Project Lab missing"),
    ("Investment decisions", "IRR", "0 = sum(C_t/(1+IRR)^t)", "Yes", "Yes", "Yes", "Yes", "No", "No", "Later", "Implemented/tested; Project Lab missing"),
    ("Investment decisions", "Profitability index", "PI = NPV / Initial investment", "Yes", "Yes", "Yes", "No", "No", "No", "Later", "Project Lab"),
    ("Stocks", "Basic stock valuation", "P_0 = sum(DIV_t/(1+r)^t)", "Yes", "Yes", "Yes", "Yes", "No", "No", "Later", "DDM tests exist; not current product"),
    ("Stocks", "Constant growth stock", "P_0 = DIV_1 / (r-g)", "Yes", "Yes", "Yes", "Yes", "No", "No", "Later", "DDM tests; not current product"),
    ("Derivatives", "Forward price continuous compounding", "F = S x exp((r + storage - yield) x T)", "Yes", "Yes", "Yes", "No", "No", "No", "Only if scope expands", "Derivatives out of scope"),
    ("Derivatives", "Put-call parity", "C + PV(K) = P + S_0", "Yes", "Yes", "Yes", "No", "No", "No", "Only if scope expands", "Derivatives out of scope"),
    ("Derivatives", "Option payoff", "Call=max(S_T-K,0); Put=max(K-S_T,0)", "Yes", "Yes", "Yes", "No", "No", "No", "Only if scope expands", "Derivatives out of scope"),
    ("Derivatives", "Binomial option pricing", "Option value = discounted expected risk-neutral payoff", "Yes", "Yes", "Yes", "No", "No", "No", "Only if scope expands", "Derivatives out of scope"),
    ("Derivatives", "Black-Scholes", "C = S_0N(d1) - Ke^(-rT)N(d2)", "Yes", "Yes", "Yes", "No", "No", "No", "Only if scope expands", "Derivatives out of scope"),
]


EXAM_ROWS = [
    ("SampleExam1", "Problem 1", "Perfect capital markets, WACC, capital structure theory", "MM, WACC, project discount-rate reasoning", "No", "Partial", "No", "Corporate finance", "Theory MCQs; not ideal as product output"),
    ("SampleExam1", "Problem 2", "Investment project FCF, WACC from comparables, PV tax shield", "FCF, Hamada beta, CAPM, PVTS", "Yes", "Yes", "Partial", "Corporate finance", "Implemented as formula tests; project lab not productized"),
    ("SampleExam1", "Problem 3", "NPV vs IRR and mutually exclusive projects", "NPV, IRR, incremental IRR", "Yes", "Yes", "No", "Corporate finance", "Tests exist; project decision screen missing"),
    ("SampleExam1", "Problem 4", "Derivatives", "Derivative pricing/payoff reasoning", "No", "No", "No", "Out of scope", "Current OpenQuant does not cover course second part"),
    ("SampleExam2", "Problem 1", "Warm-up TVM, growing annuity, compounding, bonds/rates", "Compounding, growing annuity, zero-coupon, forward rates, coupon bond", "Partial", "Partial", "No", "Mixed", "Growing annuity tested; bonds/rates not productized"),
    ("SampleExam2", "Problem 2", "Multi-stage stock valuation", "DDM, terminal price, NPV of dividends, capital gain", "Yes", "Yes", "No", "Equity valuation", "Implemented as tests; current product uses FCF DCF"),
    ("SampleExam2", "Problem 3", "Beta from correlation and CAPM", "beta = rho sigma_i / sigma_m; CAPM", "Yes", "Yes", "Partial", "Risk/return", "CAPM productized through WACC; correlation-beta is test-only"),
    ("SampleExam2", "Problem 4", "Asset allocation and portfolio volatility", "Portfolio return, variance, min variance weight", "Yes", "Yes", "No", "Portfolio", "Tests exist; portfolio lab missing"),
    ("SampleExam2", "Problem 5", "Idiosyncratic variance and Sharpe ratio", "Idiosyncratic variance, Sharpe", "Yes", "Yes", "No", "Portfolio", "Tests exist; portfolio lab missing"),
    ("SampleExam3", "Problem 1", "PV comparison of business-sale offers", "PV, NPV", "No", "Yes", "No", "Time value", "Could be demonstration; not needed in current product"),
    ("SampleExam3", "Problem 2", "Retirement savings and withdrawals", "Growing annuity, annuity payout", "No", "Partial", "No", "Time value", "Could be demonstration; not current product"),
    ("SampleExam3", "Problem 3", "Pricing of risk", "Risk premium / CAPM-style reasoning", "No", "Partial", "Partial", "Risk/return", "Needs detailed mapping if used"),
    ("SampleExam3", "Problem 4", "Asset allocation", "Portfolio return/risk", "No", "Partial", "No", "Portfolio", "Future Portfolio Lab"),
    ("SampleExam3", "Problem 5", "Free cash flows and NPV", "FCF, NPV", "No", "Partial", "Partial", "Corporate finance", "Useful for Project Lab traceability"),
]


OPENQUANT_ROWS = [
    ("api/main.py", "Analyse endpoint", "Orchestrates ticker validation, data fetch, FCF, WACC, DCF, reverse DCF, sensitivity, multiples", "Yes", "Yes", "Yes", "Current core product endpoint"),
    ("api/main.py", "Calibration endpoint", "Exposes backtest/results/calibration_summary.json", "Yes", "Yes", "Yes", "Main results only; no Markdown report"),
    ("core/data.py", "Real data layer", "SEC EDGAR statements, yfinance prices, validation, caching", "Yes", "Yes", "Yes", "US-listed companies only"),
    ("core/fcf.py", "FCF analysis", "Free cash flow history, growth, margins, warnings", "Yes", "Yes", "Yes", "Core current product"),
    ("core/wacc.py", "Risk and discount rate", "Beta, CAPM cost of equity, cost of debt, WACC, Hamada helpers", "Yes", "Yes", "Yes", "Core current product"),
    ("core/dcf.py", "Forward DCF", "Scenario valuation, terminal value, NPV, IRR, PV tax shield helpers", "Yes", "Yes", "Yes", "DCF productized; NPV/IRR/PVTS mostly test/support"),
    ("core/reverse_dcf.py", "Market-implied expectations", "Solve implied FCF growth required by current price", "Yes", "Yes", "Yes", "Central product differentiator"),
    ("core/sensitivity.py", "Sensitivity analysis", "Growth x WACC and terminal growth x WACC tables", "Yes", "Yes", "Yes", "Productized"),
    ("core/multiples.py", "Market cross-check", "P/E, EV/EBITDA, FCF yield", "Yes", "Yes", "Yes", "Not course core, but useful sanity check"),
    ("core/suitability.py", "Suitability gate", "DCF appropriateness checks and alternatives", "Yes", "Yes", "Yes", "Important for not overclaiming"),
    ("core/assumption_diagnostic.py", "Diagnostic scoring", "Assumption risk scoring", "Yes", "Yes", "Yes", "Wired into /analyse and shown in the model reliability panel"),
    ("core/red_flags.py", "Red flag summary", "Aggregates diagnostic and suitability caveats", "Yes", "Yes", "Yes", "Wired into /analyse and shown before detailed formula evidence"),
    ("core/audit_trail.py", "Formula trace", "Display formula references and inputs used", "Yes", "Yes", "No", "Wired into /analyse and UI; add dedicated tests next"),
    ("backtest/", "Historical validation", "Run OpenQuant as of past dates and compare later outcomes", "Yes", "Yes", "Yes", "Backtest informativeness, not formula correctness"),
    ("tests/test_epfl_exam1.py", "Sample exam 1 formula tests", "FCF, Hamada, CAPM, PVTS, NPV, IRR", "Yes", "No", "Yes", "Formula correctness layer"),
    ("tests/test_epfl_exam2.py", "Sample exam 2 formula tests", "Growing annuity, DDM, beta/CAPM, portfolio, Sharpe", "Yes", "No", "Yes", "Formula correctness layer"),
    ("tests/test_backtest_analysis.py", "Backtest metric tests", "Locks headline backtest metrics and API payload", "Yes", "Yes", "Yes", "Prevents silent drift in main reliability result"),
    ("frontend/src", "Stock Valuation Lab UI", "Ticker analysis, scenarios, assumptions, reliability panel", "Yes", "Yes", "Yes", "Next improvement: make market-implied belief even more central"),
]


MISSING_ROWS = [
    ("Product gap", "Project Decision Lab", "Not productized", "High later", "Needed to directly express value enhancement through project decisions", "Build NPV/IRR/PI workflow with real or user-entered project cash flows"),
    ("Product gap", "Bonds & Rates Lab", "Not productized", "Medium later", "Course/workbook/exams include bonds, spot/forward rates, duration", "Build bond price, YTM, duration, rate-shock demo"),
    ("Product gap", "Portfolio Lab", "Not productized", "Medium later", "Course/workbook/exams include diversification, min variance, Sharpe", "Build portfolio risk/return and diversification view"),
    ("Product gap", "Capital Structure Policy Lab", "Not productized", "Medium later", "Course first part includes policies/strategies maximizing firm value", "Build WACC/APV/debt tax shield/distress trade-off workflow"),
    ("Out of scope", "Derivatives Lab", "Out of scope", "Only if scope expands", "Course second part exists but current OpenQuant does not cover it", "Only after corporate-finance product is solid"),
    ("Traceability gap", "SampleExam3 detailed mapping", "Partial", "Medium", "Not currently represented by dedicated tests", "Add tests only for concepts relevant to planned modules"),
    ("Product gap", "NWC component explanation", "Partial", "High later", "Important for FCF and project decisions", "Expose AR/inventory/AP and delta NWC in FCF audit"),
    ("Product gap", "APV full workflow", "Partial", "Medium later", "PV tax shield exists but APV is not a product flow", "Build APV as alternative to WACC valuation when leverage changes"),
    ("Product gap", "Backtest breadth", "Partial", "High later", "Current headline backtest is one 2014-2024 stock valuation run", "Add multiple as-of dates and scenario variants"),
    ("Testing gap", "Audit trail dedicated tests", "Productized, not directly tested", "Medium", "Audit trail is now returned by API/UI but lacks focused tests", "Add tests for audit payload shape and formula references"),
    ("UX gap", "Market-implied belief as first screen", "Partial", "High", "README/strategy now say this, UI still includes fair-value framing", "Move reverse DCF and assumptions before intrinsic-value verdict language"),
]


def setup_sheet(ws, title, subtitle=None):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 24
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(italic=True, color="666666")
        ws["A2"].alignment = Alignment(wrap_text=True)


def write_table(ws, start_row, headers, rows):
    thin = Side(style="thin", color="D9E2F3")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(start_row, col, header)
        cell.font = Font(bold=True, color="1F1F1F")
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = Border(top=thin, bottom=thin)
    for r_idx, row in enumerate(rows, start_row + 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(r_idx, c_idx, val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(bottom=Side(style="hair", color="E7E6E6"))
            if val in {"Yes", "Productized"}:
                cell.fill = PatternFill("solid", fgColor="E2F0D9")
            elif val in {"No", "Out of scope"}:
                cell.fill = PatternFill("solid", fgColor="FCE4D6")
            elif val in {"Partial", "Later", "Only if scope expands", "Maybe"}:
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{start_row + len(rows)}"


def set_widths(ws, widths):
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def add_project_plan(wb):
    ws = wb.active
    ws.title = "Project Plan"
    setup_sheet(ws, "OpenQuant Project Plan", "Objectives, scope, source material, and current product definition.")
    rows = [
        ("Course objective", COURSE_OBJECTIVE),
        ("Course source", "EPFL Principles of Finance; likely built from Berk & DeMarzo Corporate Finance, Pearson, 2nd Ed., 2010."),
        ("OpenQuant goal", "Practical application of the course objective using real listed-company data."),
        ("Current product", "Stock Valuation Lab: what must be true for the current market price to make sense?"),
        ("Core output", "Real-data valuation insight supported by assumptions, formulas, suitability checks, and reliability evidence."),
        ("Backtest role", "Historical authenticity layer: past OpenQuant conclusions are compared with later observed data."),
        ("Formula role", "Formulas are implementation tools and evidence, not the final product."),
        ("Current out of scope", "Derivatives; bonds/rates lab; portfolio lab; project decision lab; full capital-structure policy lab."),
        ("Next product direction", "Make market-implied assumptions the first-class output; keep fair value as a derived view."),
    ]
    write_table(ws, 4, ["Item", "Definition"], rows)
    ws["D4"] = "Coverage Snapshot"
    ws["D4"].font = Font(bold=True)
    snapshot = [
        ("Concept rows", "=COUNTA('Concept Inventory'!B5:B200)"),
        ("Formula rows", "=COUNTA('Formula Inventory'!B5:B200)"),
        ("Productized concepts", "=COUNTIF('Concept Inventory'!H5:H200,\"Yes\")"),
        ("Productized formulas", "=COUNTIF('Formula Inventory'!H5:H200,\"Yes\")"),
        ("Missing / later items", "=COUNTA('Missing Coverage'!B5:B200)"),
    ]
    for i, (label, formula) in enumerate(snapshot, 5):
        ws.cell(i, 4, label)
        ws.cell(i, 5, formula)
        ws.cell(i, 5).number_format = "0"
    set_widths(ws, [24, 92, 3, 28, 18])


def build():
    wb = Workbook()
    add_project_plan(wb)

    ws = wb.create_sheet("Concept Inventory")
    setup_sheet(ws, "Concept Inventory", "Course notions and competencies mapped to OpenQuant implementation and product scope.")
    write_table(ws, 4, ["Course block", "Concept / competency", "Why it matters", "Source", "Exam tested?", "Workbook?", "OpenQuant implemented?", "OpenQuant productized?", "Should productize?", "Notes"], CONCEPT_ROWS)
    set_widths(ws, [20, 34, 34, 28, 14, 12, 18, 18, 18, 48])

    ws = wb.create_sheet("Formula Inventory")
    setup_sheet(ws, "Formula Inventory", "Grouped formula inventory from formula sheet, workbook, exams, and OpenQuant implementation.")
    write_table(ws, 4, ["Area", "Formula / method", "Formula", "Formula sheet?", "Workbook?", "Sample exams?", "OpenQuant implemented?", "OpenQuant productized?", "Related to current product?", "Should productize?", "Notes"], FORMULA_ROWS)
    set_widths(ws, [18, 32, 54, 14, 12, 14, 20, 20, 22, 18, 44])

    ws = wb.create_sheet("Exam Mapping")
    setup_sheet(ws, "Exam Mapping", "Sample exam knowledge mapped to implemented tests and product scope.")
    write_table(ws, 4, ["Exam", "Problem", "Knowledge tested", "Formula / method", "OpenQuant test?", "OpenQuant implemented?", "OpenQuant productized?", "Scope", "Notes"], EXAM_ROWS)
    set_widths(ws, [16, 16, 42, 38, 18, 20, 20, 20, 48])

    ws = wb.create_sheet("OpenQuant Coverage")
    setup_sheet(ws, "OpenQuant Coverage", "Repo-level mapping of product capabilities, formula support, and tests.")
    write_table(ws, 4, ["Module / file", "Capability", "What it covers", "Formula implemented?", "Productized?", "Tested?", "Notes"], OPENQUANT_ROWS)
    set_widths(ws, [30, 30, 58, 20, 16, 14, 48])

    ws = wb.create_sheet("Source Workbook")
    setup_sheet(ws, "Source Workbook", "EPFL master workbook sheets and formula-cell counts, used as source inventory.")
    write_table(ws, 4, ["Workbook sheet", "Formula cells", "What it covers"], WORKBOOK_SHEETS)
    set_widths(ws, [28, 16, 80])

    ws = wb.create_sheet("Missing Coverage")
    setup_sheet(ws, "Missing Coverage", "Items outside current product scope or not yet fully covered.")
    write_table(ws, 4, ["Type", "Item", "Status today", "Suggested priority", "Why it matters / why missing", "To make full"], MISSING_ROWS)
    set_widths(ws, [18, 32, 18, 18, 58, 58])

    ws = wb.create_sheet("Sources")
    setup_sheet(ws, "Sources", "Source material used to build this project plan workbook.")
    write_table(ws, 4, ["Source type", "File / location", "Used for"], SOURCE_FILES)
    set_widths(ws, [20, 52, 72])

    ws = wb.create_sheet("Checks")
    setup_sheet(ws, "Checks", "Basic workbook integrity checks.")
    checks = [
        ("Concept inventory has rows", "=COUNTA('Concept Inventory'!B5:B200)", ">0", '=IF(B5>0,"OK","CHECK")', "Should be populated"),
        ("Formula inventory has rows", "=COUNTA('Formula Inventory'!B5:B250)", ">0", '=IF(B6>0,"OK","CHECK")', "Should be populated"),
        ("Exam mapping has rows", "=COUNTA('Exam Mapping'!C5:C100)", ">0", '=IF(B7>0,"OK","CHECK")', "Should be populated"),
        ("OpenQuant coverage has rows", "=COUNTA('OpenQuant Coverage'!B5:B100)", ">0", '=IF(B8>0,"OK","CHECK")', "Should be populated"),
        ("Objective present", '=IF(ISNUMBER(SEARCH("financial valuation",\'Project Plan\'!B5)),1,0)', "1", '=IF(B9=1,"OK","CHECK")', "Official objective should appear in Project Plan"),
    ]
    write_table(ws, 4, ["Check", "Actual", "Expected", "Status", "Notes"], checks)
    set_widths(ws, [34, 22, 18, 16, 52])

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                font = copy(cell.font)
                font.name = "Aptos"
                font.sz = 10
                cell.font = font
        sheet["A1"].font = Font(name="Aptos", bold=True, size=16, color="FFFFFF")
        for row in range(1, sheet.max_row + 1):
            sheet.row_dimensions[row].height = max(sheet.row_dimensions[row].height or 15, 18)

    wb.save(OUT)


if __name__ == "__main__":
    build()
